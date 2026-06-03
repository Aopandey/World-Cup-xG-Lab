from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = PROJECT_ROOT / "data" / "squads" / "processed" / "world_cup_2026_squads.csv"
OUTPUT_PATH = (
    PROJECT_ROOT
    / "data"
    / "squads"
    / "processed"
    / "world_cup_2026_players_fbref_scrape_handoff.xlsx"
)

POSITION_ORDER = {
    "Goalkeeper": 0,
    "Defender": 1,
    "Midfielder": 2,
    "Forward": 3,
    "Unknown": 4,
}


def build_handoff_table(squads: pd.DataFrame) -> pd.DataFrame:
    """Return the player-level handoff table for FBref scraping."""
    confirmed = squads[squads["squad_status"].eq("confirmed")].copy()
    confirmed["_position_order"] = confirmed["position_group"].map(POSITION_ORDER).fillna(99)
    confirmed = confirmed.sort_values(
        ["world_cup_team", "_position_order", "player"],
    ).reset_index(drop=True)

    return confirmed.rename(
        columns={
            "player": "Player",
            "world_cup_team": "National Team",
            "league": "League",
            "club": "Club Team",
            "position": "Position",
        }
    )[["Player", "National Team", "League", "Club Team", "Position"]]


def build_league_counts(handoff: pd.DataFrame) -> pd.DataFrame:
    """Summarize player counts by league."""
    return (
        handoff.assign(League=handoff["League"].fillna("Unknown / verify manually"))
        .groupby("League", dropna=False)
        .agg(Players=("Player", "count"), National_Teams=("National Team", "nunique"))
        .reset_index()
        .sort_values(["Players", "League"], ascending=[False, True])
    )


def build_team_counts(handoff: pd.DataFrame) -> pd.DataFrame:
    """Summarize player counts by national team."""
    return (
        handoff.groupby("National Team")
        .agg(
            Players=("Player", "count"),
            Leagues=("League", "nunique"),
            Clubs=("Club Team", "nunique"),
        )
        .reset_index()
        .sort_values("National Team")
    )


def build_data_dictionary() -> pd.DataFrame:
    """Return a short data dictionary for the handoff workbook."""
    return pd.DataFrame(
        {
            "Field": ["Player", "National Team", "League", "Club Team", "Position"],
            "Description": [
                "Confirmed 2026 World Cup squad player name from the project squad table.",
                "Normalized national team name used by the dashboard.",
                "Current club league label from the squad source/manual override; verify Unknown rows manually before scraping.",
                "Current club/team listed for the player.",
                "Dashboard position group: Goalkeeper, Defender, Midfielder, or Forward.",
            ],
        }
    )


def style_workbook(path: Path) -> None:
    """Apply simple workbook formatting for easier manual review."""
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="0B222D")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            adjusted_width = min(max(max_length + 2, 12), 45)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
                adjusted_width
            )

    workbook.save(path)


def main() -> None:
    """Export a World Cup squad player workbook for FBref scraping."""
    if not INPUT_PATH.exists():
        raise FileNotFoundError(
            f"Processed squad file not found: {INPUT_PATH}. "
            "Run python scripts/apply_manual_squad_overrides.py first."
        )

    squads = pd.read_csv(INPUT_PATH)
    handoff = build_handoff_table(squads)
    league_counts = build_league_counts(handoff)
    team_counts = build_team_counts(handoff)
    data_dictionary = build_data_dictionary()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_PATH

    try:
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
    except PermissionError:
        output_path = OUTPUT_PATH.with_name(
            f"{OUTPUT_PATH.stem}_updated{OUTPUT_PATH.suffix}"
        )
        writer = pd.ExcelWriter(output_path, engine="openpyxl")

    with writer:
        handoff.to_excel(writer, sheet_name="fbref_scrape_players", index=False)
        league_counts.to_excel(writer, sheet_name="league_counts", index=False)
        team_counts.to_excel(writer, sheet_name="team_counts", index=False)
        data_dictionary.to_excel(writer, sheet_name="data_dictionary", index=False)

    style_workbook(output_path)

    print(f"Saved: {output_path}")
    print(f"Rows: {len(handoff):,}")
    print(f"National teams: {handoff['National Team'].nunique():,}")
    print(f"Unique leagues: {handoff['League'].nunique(dropna=True):,}")


if __name__ == "__main__":
    main()
